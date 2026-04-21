"""
Video Judge Excel workbook builder (2026-04-20).

Produces one .xlsx workbook with a sheet per event, long-format rows
(row per competitor per run; row per partnered pair per run) with two
video-judge timer/score columns per row. Used during show-prep so a
video-judging team can record independent times off recorded heats.

Rules agreed in plan-eng-review (see docs/VIDEO_JUDGE_BRACKET_PLAN.md):
  - Skip events where scoring_type == 'bracket' (Birling).
  - Skip events whose normalised name is in config.LIST_ONLY_EVENT_NAMES
    (Axe Throw, Peavey Log Roll, Caber Toss, Pulp Toss — sign-up events
    with no heats).
  - Dual-run day-split events (Chokerman's Race, Speed Climb) emit TWO
    sheets: "Event Name - Run 1" and "Event Name - Run 2".
  - Triple-run events (Axe Throw via requires_triple_runs) emit one
    sheet with three stacked rows per competitor (Run 1 / 2 / 3).
  - Partnered events emit one row per pair per run ("Alice & Bob").
  - Hits/score events use "VJ Score 1 / VJ Score 2" column headers;
    timed events use "VJ Timer 1 / VJ Timer 2".
  - Sheet names: truncate to 31 chars (openpyxl limit), strip Excel-
    invalid characters [ ] : * ? / \\, and dedupe collisions by
    appending " (2)", " (3)", etc.
  - Stable row order per sheet: heat_number ASC, run_number ASC,
    stand_number ASC, competitor_name ASC.
"""

from __future__ import annotations

import re
from collections import OrderedDict

from config import LIST_ONLY_EVENT_NAMES
from models import Event, Heat, Tournament
from models.competitor import CollegeCompetitor, ProCompetitor
from services.partner_resolver import pair_competitors_for_heat

# Events whose run 1 and run 2 live on different days — emit as two sheets so
# the Friday video-judge crew and the Saturday crew each get a clean tab.
from config import DAY_SPLIT_EVENT_NAMES

# Excel sheet name constraints (openpyxl)
_SHEET_NAME_INVALID = re.compile(r"[\[\]\:\*\?/\\]")
_SHEET_NAME_MAX = 31


class VideoJudgeWorkbookError(RuntimeError):
    """Raised when openpyxl cannot write the workbook (invalid sheet name,
    filesystem issue, etc.). The route catches this and flashes a user-
    friendly message instead of a 500."""


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _normalise_event_name(name: str) -> str:
    return "".join(ch for ch in (name or "").lower() if ch.isalnum())


def _is_list_only(event: Event) -> bool:
    return _normalise_event_name(event.name) in LIST_ONLY_EVENT_NAMES


def _num_runs(event: Event) -> int:
    if getattr(event, "requires_triple_runs", False):
        return 3
    if getattr(event, "requires_dual_runs", False):
        return 2
    return 1


def _column_labels(event: Event) -> tuple[str, str]:
    """Return (col1, col2) labels for the two VJ columns on this event."""
    if event.scoring_type in ("time", "distance"):
        return "VJ Timer 1", "VJ Timer 2"
    return "VJ Score 1", "VJ Score 2"


def _sanitize_sheet_name(name: str) -> str:
    """Strip Excel-invalid chars and truncate to 31 chars."""
    clean = _SHEET_NAME_INVALID.sub("", name or "").strip()
    if not clean:
        clean = "Sheet"
    return clean[:_SHEET_NAME_MAX]


def _dedupe_sheet_names(names: list[str]) -> list[str]:
    """Given a list of sanitised sheet names that may collide after truncation,
    append ' (2)', ' (3)', etc. to duplicates. Preserves input order."""
    seen: dict[str, int] = {}
    out: list[str] = []
    for n in names:
        if n not in seen:
            seen[n] = 1
            out.append(n)
            continue
        seen[n] += 1
        suffix = f" ({seen[n]})"
        base_max = _SHEET_NAME_MAX - len(suffix)
        new_name = f"{n[:base_max]}{suffix}"
        # Very defensively, re-check for recursion if the truncated new_name
        # itself collides (e.g. ultra-long events).  Bump counter again.
        while new_name in out:
            seen[n] += 1
            suffix = f" ({seen[n]})"
            base_max = _SHEET_NAME_MAX - len(suffix)
            new_name = f"{n[:base_max]}{suffix}"
        out.append(new_name)
    return out


# ---------------------------------------------------------------------------
# Row building
# ---------------------------------------------------------------------------


def _load_competitor_lookup(event: Event, comp_ids: list) -> dict:
    ids = sorted({int(cid) for cid in comp_ids if cid is not None})
    if not ids:
        return {}
    model = CollegeCompetitor if event.event_type == "college" else ProCompetitor
    rows = model.query.filter(model.id.in_(ids)).all()
    return {c.id: c for c in rows}


def _team_code(event: Event, comp) -> str | None:
    if event.event_type != "college":
        return None
    team = getattr(comp, "team", None)
    return team.team_code if team else None


def _rows_for_heat(event: Event, heat: Heat, comp_lookup: dict) -> list[dict]:
    """Build row dicts for one heat. Each row becomes one Excel line per run."""
    assignments = heat.get_stand_assignments()
    pairs = pair_competitors_for_heat(event, heat.get_competitors(), comp_lookup)

    out: list[dict] = []
    for pair in pairs:
        primary_id = pair["primary_comp_id"]
        comp = pair["competitor"]
        stand = assignments.get(str(primary_id))
        try:
            stand_int = int(stand) if stand is not None else 999
        except (TypeError, ValueError):
            stand_int = 999
        out.append(
            {
                "heat_number": heat.heat_number,
                "run_number": heat.run_number,
                "stand_number": stand_int,
                "stand_display": stand if stand is not None else "?",
                "competitor_name": pair["name"],
                "team_code": _team_code(event, comp) if comp else None,
            }
        )
    return out


def _sheet_key_for_heat(event: Event, heat: Heat) -> str:
    """Return the sheet title this heat should land on.

    Dual-run day-split events (Chokerman, Speed Climb) get separate sheets
    per run.  Everything else uses the event display name once and stacks
    runs as extra rows.
    """
    base = event.display_name
    is_day_split_dual = event.name in DAY_SPLIT_EVENT_NAMES and getattr(
        event, "requires_dual_runs", False
    )
    if is_day_split_dual:
        return f"{base} - Run {heat.run_number}"
    return base


def build_video_judge_rows(tournament: Tournament) -> "OrderedDict[str, dict]":
    """Walk every pro and college event for the tournament and return an
    ordered dict keyed by raw sheet title → {'event': Event, 'rows': [...]}.

    Sheet sanitisation and dedup happen at write-time so tests can assert
    on the un-truncated title too.
    """
    events = (
        Event.query.filter_by(tournament_id=tournament.id)
        .order_by(Event.event_type.asc(), Event.id.asc())
        .all()
    )

    sheets: "OrderedDict[str, dict]" = OrderedDict()

    for event in events:
        if event.scoring_type == "bracket":
            continue  # Birling — no video timing
        if _is_list_only(event):
            continue  # sign-up events have no heats

        heats = (
            Heat.query.filter_by(event_id=event.id)
            .order_by(Heat.heat_number.asc(), Heat.run_number.asc())
            .all()
        )
        if not heats:
            continue

        # Batch-load all competitors across the event's heats — avoids N+1
        # when walking heat-by-heat.
        all_comp_ids: list = []
        for heat in heats:
            all_comp_ids.extend(heat.get_competitors())
        comp_lookup = _load_competitor_lookup(event, all_comp_ids)

        for heat in heats:
            sheet_title = _sheet_key_for_heat(event, heat)
            entry = sheets.setdefault(sheet_title, {"event": event, "rows": []})
            entry["rows"].extend(_rows_for_heat(event, heat, comp_lookup))

    # Stable sort per sheet.  Row order: heat_number, run_number, stand, name.
    for entry in sheets.values():
        entry["rows"].sort(
            key=lambda r: (
                r["heat_number"],
                r["run_number"],
                r["stand_number"],
                r["competitor_name"].lower(),
            )
        )

    return sheets


# ---------------------------------------------------------------------------
# Workbook writer
# ---------------------------------------------------------------------------


def write_workbook(sheets: "OrderedDict[str, dict]", path: str) -> None:
    """Write the sheets dict to an xlsx file.

    Raises VideoJudgeWorkbookError on any openpyxl / filesystem failure so
    the calling route can flash a user-friendly message.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    # Workbook() ships with one default sheet; remove it so our first
    # tab is at index 0 with the right title.
    default = wb.active
    wb.remove(default)

    if not sheets:
        # Empty workbook: still produce a placeholder sheet so the download
        # isn't corrupt, and the user sees a clear "no events yet" signal.
        ws = wb.create_sheet("No Events")
        ws["A1"] = "This tournament has no events with heats yet."
        try:
            wb.save(path)
        except Exception as exc:  # noqa: BLE001
            raise VideoJudgeWorkbookError(f"Could not save workbook: {exc}") from exc
        return

    # Sanitise + dedupe sheet titles first so we know upfront what to write.
    raw_titles = list(sheets.keys())
    sanitised = [_sanitize_sheet_name(t) for t in raw_titles]
    final_titles = _dedupe_sheet_names(sanitised)

    for title, raw_title in zip(final_titles, raw_titles):
        entry = sheets[raw_title]
        event = entry["event"]
        rows = entry["rows"]
        ws = wb.create_sheet(title)
        col1, col2 = _column_labels(event)

        headers = [
            "Heat",
            "Run",
            "Stand",
            "Competitor",
            "Team" if event.event_type == "college" else "",
            col1,
            col2,
            "Status",
            "Reason",
        ]
        # Drop the empty "Team" column slot for pro events.
        if event.event_type != "college":
            headers = [h for h in headers if h != ""]

        ws.append(headers)
        # Header styling: bold + light grey fill so the VJ crew can scan.
        header_fill = PatternFill("solid", fgColor="E8E8E8")
        bold = Font(bold=True)
        for cell in ws[1]:
            cell.font = bold
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        n_runs = _num_runs(event)
        for row in rows:
            base_cols = [
                row["heat_number"],
                row["run_number"],
                row["stand_display"],
                row["competitor_name"],
            ]
            if event.event_type == "college":
                base_cols.append(row["team_code"] or "")
            # Triple-run: stack run 1 / 2 / 3 as separate lines per competitor.
            if n_runs == 3:
                for run_idx in range(1, 4):
                    ws.append(
                        [
                            row["heat_number"],
                            run_idx,
                            row["stand_display"],
                            row["competitor_name"],
                            *(
                                [row["team_code"] or ""]
                                if event.event_type == "college"
                                else []
                            ),
                            "",  # VJ 1
                            "",  # VJ 2
                            "",  # status
                            "",  # reason
                        ]
                    )
            else:
                ws.append(
                    [
                        *base_cols,
                        "",  # VJ 1
                        "",  # VJ 2
                        "",  # status
                        "",  # reason
                    ]
                )

        # Column widths sized for printing on a Letter-landscape screen.
        widths = [6, 5, 7, 28]
        if event.event_type == "college":
            widths.append(8)
        widths += [12, 12, 10, 22]
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[chr(ord("A") + idx - 1)].width = width

    try:
        wb.save(path)
    except Exception as exc:  # noqa: BLE001
        raise VideoJudgeWorkbookError(f"Could not save workbook: {exc}") from exc
