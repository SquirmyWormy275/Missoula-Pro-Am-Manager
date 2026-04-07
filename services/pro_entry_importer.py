"""
Pro competitor entry importer for Google Forms xlsx exports.

Reads the first sheet of an xlsx workbook (Google always puts "Form Responses 1"
first) and returns a list of parsed entry dicts ready for review and DB commit.
"""
from datetime import datetime

import openpyxl

from services.gear_sharing import infer_equipment_categories

# Waiver column is identified by this prefix (full text is too long to quote here)
_WAIVER_HEADER_START = 'I know that logging events'

# Maps stripped form header -> (canonical event name, fee amount)
_EVENT_MAP = {
    'Springboard (L)':                  ('Springboard',              10),
    'Springboard (R)':                  ('Springboard',              10),
    'Intermediate 1-Board Springboard': ('Pro 1-Board',             10),
    '1-Board Springboard':              ('Pro 1-Board',             10),
    'Pro 1-Board':                      ('Pro 1-Board',             10),
    "Men's Underhand":                  ("Men's Underhand",          10),
    "Women's Underhand":                ("Women's Underhand",        10),
    "Women's Standing Block":           ("Women's Standing Block",   10),
    "Men's Standing Block":             ("Men's Standing Block",     10),
    "Men's Single Buck":                ("Men's Single Buck",         5),
    "Women's Single Buck":              ("Women's Single Buck",       5),
    "Men's Double Buck":                ("Men's Double Buck",         5),
    "Women's Double Buck":              ("Women's Double Buck",       5),
    'Double Buck':                      ("Men's Double Buck",         5),
    'Jack & Jill':                      ('Jack & Jill Sawing',        5),
    'Jack & Jill Sawing':               ('Jack & Jill Sawing',        5),
    'Jack Jill':                        ('Jack & Jill Sawing',        5),
    'Hot Saw':                          ('Hot Saw',                   5),
    'Obstacle Pole':                    ('Obstacle Pole',             5),
    'Speed Climb':                      ('Pole Climb',                5),
    'Pole Climb':                       ('Pole Climb',                5),
    'Cookie Stack':                     ('Cookie Stack',              5),
    '3-Board Jigger':                   ('3-Board Jigger',            5),
    '3 Board Jigger':                   ('3-Board Jigger',            5),
    'Partnered Axe Throw':              ('Partnered Axe Throw',       5),
    'Axe Throw':                        ('Partnered Axe Throw',       5),
    "Men's Stock Saw":                  ("Men's Stock Saw",           5),
    "Women's Stock Saw":                ("Women's Stock Saw",         5),
    'Stock Saw':                        ("Men's Stock Saw",           5),
}

# Maps lowercased stripped partner-column header -> canonical event name
_PARTNER_COLS = {
    "men's double buck partner name": "Men's Double Buck",
    "jack & jill partner name":       "Jack & Jill Sawing",
    "partnered axe throw 2":          "Partnered Axe Throw",
}

_TRUE_MARKERS = {'yes', 'y', 'true', '1', 'x'}


def _find_column_index(stripped_headers: list[str], candidates: list[str]) -> int | None:
    """Find a header index by exact or contains-match against normalized candidates."""
    lowered = [str(h or '').strip().lower() for h in stripped_headers]
    normalized_candidates = [c.strip().lower() for c in candidates if c and c.strip()]

    for candidate in normalized_candidates:
        if candidate in lowered:
            return lowered.index(candidate)
    for idx, header in enumerate(lowered):
        if any(candidate in header for candidate in normalized_candidates):
            return idx
    return None


def _yes(val) -> bool:
    """Return True only when value is the string 'Yes' (case-insensitive)."""
    return str(val).strip().lower() == 'yes' if val is not None else False


def _get(row: tuple, col) -> object:
    """Safely retrieve a value from a row tuple by column index."""
    if col is None or col >= len(row):
        return None
    return row[col]


def parse_pro_entries(filepath: str) -> list:
    """
    Parse a Google Forms xlsx export (first sheet) and return a list of dicts.

    Each dict contains all form data needed for review and DB import.
    Datetime objects are converted to ISO strings for JSON serialisability.
    Rows where 'Full Name' is blank are silently skipped.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.worksheets[0]

    # Read and strip every header cell from row 1
    raw_headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    stripped = [h.strip() if isinstance(h, str) else (h or '') for h in raw_headers]

    # Exact-match lookup: stripped_header -> 0-based column index.
    # Also index the first line for prompts that include explanatory line breaks.
    hmap = {}
    for i, h in enumerate(stripped):
        if not h:
            continue
        hmap[h] = i
        first_line = h.splitlines()[0].strip()
        if first_line:
            hmap.setdefault(first_line, i)

    # Prefix-matched special columns
    waiver_col      = next((i for i, h in enumerate(stripped) if h.startswith(_WAIVER_HEADER_START)), None)
    gear_detail_col = next((i for i, h in enumerate(stripped) if h.lower().startswith('if yes, provide')), None)
    notes_col       = next((i for i, h in enumerate(stripped) if h.lower().startswith('anything else we should know')), None)
    slow_heat_col   = _find_column_index(stripped, [
        'springboard slow heat',
        'slow heat springboard',
        'relegated to slow heat',
        'springboard slow',
    ])

    entries = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Skip rows with no name (trailing blank rows, etc.)
        name_val = _get(row, hmap.get('Full Name'))
        if not name_val or not str(name_val).strip():
            continue

        # --- Timestamp ---
        ts = _get(row, hmap.get('Timestamp'))
        timestamp_str = ts.isoformat() if isinstance(ts, datetime) else (str(ts) if ts else None)

        # --- Basic personal fields ---
        email = str(_get(row, hmap.get('Email Address')) or '').strip() or None
        name  = str(name_val).strip()

        gender_raw = str(_get(row, hmap.get('Gender')) or '').strip()
        if gender_raw == 'Male':
            gender = 'M'
        elif gender_raw == 'Female':
            gender = 'F'
        else:
            gender = gender_raw[:1].upper() if gender_raw else ''

        mailing_address = str(_get(row, hmap.get('Mailing Address')) or '').strip() or None

        # Phone: Google exports as float (e.g. 3033496312.0) -> int -> str
        phone_val = _get(row, hmap.get('Phone Number'))
        if phone_val is not None:
            try:
                phone = str(int(float(phone_val)))
            except (TypeError, ValueError):
                phone = str(phone_val).strip() or None
        else:
            phone = None

        ala_member = _yes(_get(row, hmap.get('Are you a current ALA member?')))

        # --- Events and fees ---
        events        = []
        chopping_fees = 0
        other_fees    = 0

        for form_header, (event_name, fee) in _EVENT_MAP.items():
            if _yes(_get(row, hmap.get(form_header))):
                events.append(event_name)
                if fee == 10:
                    chopping_fees += 10
                else:
                    other_fees += 5

        # --- Relay lottery ---
        relay_lottery = _yes(_get(row, hmap.get('I would like to enter into the Pro-Am lottery')))
        relay_fee     = 5 if relay_lottery else 0
        total_fees    = chopping_fees + other_fees + relay_fee

        # --- Partners (by lowercased stripped header match) ---
        partners = {}
        for stripped_header, event_name in _PARTNER_COLS.items():
            col_idx = next(
                (i for i, h in enumerate(stripped) if h.lower() == stripped_header),
                None
            )
            if col_idx is not None:
                val = _get(row, col_idx)
                if val and str(val).strip():
                    partners[event_name] = str(val).strip()

        # --- Gear sharing ---
        gear_sharing = _yes(_get(row, hmap.get('Are you sharing gear?')))
        gd_val = _get(row, gear_detail_col) if gear_detail_col is not None else None
        gear_sharing_details = str(gd_val).strip() if gd_val and str(gd_val).strip() else None

        # --- Waiver ---
        waiver_val = _get(row, waiver_col) if waiver_col is not None else None
        if waiver_val:
            wv = str(waiver_val).strip()
            waiver_accepted = wv == 'Yes' or wv.startswith('I know')
        else:
            waiver_accepted = False

        sig_val = _get(row, hmap.get('Signature'))
        waiver_signature = str(sig_val).strip() if sig_val and str(sig_val).strip() else None

        # --- Notes ---
        nv = _get(row, notes_col) if notes_col is not None else None
        notes = str(nv).strip() if nv and str(nv).strip() else None
        slow_heat_val = _get(row, slow_heat_col) if slow_heat_col is not None else None
        springboard_slow_heat = str(slow_heat_val or '').strip().lower() in _TRUE_MARKERS

        entries.append({
            'submission_timestamp': timestamp_str,
            'email':                email,
            'name':                 name,
            'gender':               gender,
            'mailing_address':      mailing_address,
            'phone':                phone,
            'ala_member':           ala_member,
            'events':               events,
            'relay_lottery':        relay_lottery,
            'partners':             partners,
            'gear_sharing':         gear_sharing,
            'gear_sharing_details': gear_sharing_details,
            'waiver_accepted':      waiver_accepted,
            'waiver_signature':     waiver_signature,
            'notes':                notes,
            'springboard_slow_heat': springboard_slow_heat,
            'chopping_fees':        chopping_fees,
            'other_fees':           other_fees,
            'relay_fee':            relay_fee,
            'total_fees':           total_fees,
        })

    return entries


def compute_review_flags(entries: list, existing_names: list = None) -> list:
    """
    Add 'flags' (list of warning strings) and 'flag_class' (Bootstrap row class)
    to each entry dict in-place.  Returns the same list for convenience.

    Rules
    -----
    - Red  (table-danger)  : waiver_accepted is False -> 'NO WAIVER'
    - Yellow (table-warning): partner name listed but not found in this batch
                              -> 'PARTNER NOT FOUND: <name>'
    - Yellow (table-warning): gear_sharing True but gear_sharing_details blank
                              -> 'GEAR SHARING DETAILS MISSING'
    - Yellow (table-warning): name closely matches existing DB competitor (#18)
                              -> 'POSSIBLE DUPLICATE OF: <name>'
    """
    import difflib
    all_names = {e['name'].strip().lower() for e in entries}

    # Build lookup of names to check for duplicates against
    check_against = list(existing_names or [])

    for entry in entries:
        flags      = []
        flag_class = ''

        if not entry['waiver_accepted']:
            flags.append('NO WAIVER')
            flag_class = 'table-danger'

        for event_name, partner_name in entry.get('partners', {}).items():
            if partner_name and partner_name.strip().lower() not in all_names:
                flags.append(f'PARTNER NOT FOUND: {partner_name}')
                if not flag_class:
                    flag_class = 'table-warning'

        if entry['gear_sharing'] and not entry['gear_sharing_details']:
            flags.append('GEAR SHARING DETAILS MISSING')
            if not flag_class:
                flag_class = 'table-warning'
        elif entry['gear_sharing'] and entry['gear_sharing_details']:
            details = str(entry['gear_sharing_details']).strip()
            has_category_signal = bool(infer_equipment_categories(details))
            likely_has_partner = len(details.split()) >= 2
            if not has_category_signal or not likely_has_partner:
                flags.append('GEAR SHARING DETAILS MAY BE AMBIGUOUS')
                if not flag_class:
                    flag_class = 'table-warning'

        # #18 — Duplicate detection
        if check_against:
            entry_name_lower = entry['name'].strip().lower()
            matches = difflib.get_close_matches(
                entry_name_lower,
                [n.strip().lower() for n in check_against],
                n=1,
                cutoff=0.85,
            )
            if matches:
                original = next(
                    (n for n in check_against if n.strip().lower() == matches[0]),
                    matches[0]
                )
                flags.append(f'POSSIBLE DUPLICATE OF: {original}')
                if not flag_class:
                    flag_class = 'table-warning'

        entry['flags']      = flags
        entry['flag_class'] = flag_class

    return entries
