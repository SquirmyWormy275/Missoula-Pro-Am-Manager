"""Regression tests for the remedial PR fixes."""

from __future__ import annotations

import io
import json
import os
import sqlite3
import tempfile

import pytest

from tests.db_test_utils import create_test_app


@pytest.fixture()
def app():
    app, _db_path = create_test_app()
    yield app


@pytest.fixture()
def admin_client(app):
    from database import db
    from models import Tournament
    from models.user import User

    with app.app_context():
        admin = User(username='admin_fix', role='admin')
        admin.set_password('password123')
        tournament = Tournament(name='Fix QA', year=2026, status='setup')
        other = Tournament(name='Fix QA Other', year=2027, status='setup')
        db.session.add_all([admin, tournament, other])
        db.session.commit()
        app.config['_FIX_ADMIN_ID'] = admin.id
        app.config['_FIX_TOURNAMENT_ID'] = tournament.id
        app.config['_FIX_OTHER_TOURNAMENT_ID'] = other.id

    client = app.test_client()
    with client.session_transaction() as sess:
        sess['_user_id'] = str(app.config['_FIX_ADMIN_ID'])
        sess['_fresh'] = True
    return client


def test_background_jobs_run_with_app_context(app):
    from services.background_jobs import get, submit

    def _job():
        from flask import current_app

        return current_app.name

    with app.app_context():
        job_id = submit('ctx-test', _job, metadata={'tournament_id': 0, 'kind': 'ctx-test'})

    for _ in range(20):
        job = get(job_id)
        if job and job['status'] in {'completed', 'failed'}:
            break

    assert job is not None
    assert job['status'] == 'completed'
    assert job['result'] == app.name


def test_generate_async_job_status_is_tournament_bound(app, admin_client):
    tournament_id = app.config['_FIX_TOURNAMENT_ID']
    other_tournament_id = app.config['_FIX_OTHER_TOURNAMENT_ID']

    response = admin_client.post(f'/scheduling/{tournament_id}/events/generate-async')
    assert response.status_code == 202
    job_id = json.loads(response.get_data(as_text=True))['job_id']

    cross = admin_client.get(f'/scheduling/{other_tournament_id}/events/job-status/{job_id}')
    assert cross.status_code == 404


def test_export_job_status_is_tournament_bound(app, admin_client):
    tournament_id = app.config['_FIX_TOURNAMENT_ID']
    other_tournament_id = app.config['_FIX_OTHER_TOURNAMENT_ID']

    response = admin_client.post(
        f'/reporting/{tournament_id}/export-results/async',
        follow_redirects=False,
    )
    assert response.status_code == 302
    location = response.headers['Location']
    job_id = location.rsplit('/', 1)[-1]

    cross = admin_client.get(f'/reporting/{other_tournament_id}/jobs/{job_id}')
    assert cross.status_code == 404


def test_friday_feature_persists_in_schedule_config(app, admin_client):
    from database import db
    from models import Event, Tournament

    tournament_id = app.config['_FIX_TOURNAMENT_ID']
    with app.app_context():
        event = Event(
            tournament_id=tournament_id,
            name='Pro 1-Board',
            event_type='pro',
            gender='M',
            scoring_type='time',
            scoring_order='lowest_wins',
            stand_type='springboard',
            max_stands=4,
            status='pending',
        )
        db.session.add(event)
        db.session.commit()
        event_id = event.id

    response = admin_client.post(
        f'/scheduling/{tournament_id}/friday-night',
        data={'event_ids': [str(event_id)], 'notes': 'Friday showcase', 'action': 'save'},
        follow_redirects=False,
    )
    assert response.status_code == 302

    with app.app_context():
        tournament = Tournament.query.get(tournament_id)
        saved = tournament.get_schedule_config()
        assert saved['friday_pro_event_ids'] == [event_id]
        assert saved['friday_feature_notes'] == 'Friday showcase'


def test_activate_competition_requires_post(app, admin_client):
    from models import Tournament

    tournament_id = app.config['_FIX_TOURNAMENT_ID']
    get_response = admin_client.get(f'/tournament/{tournament_id}/activate/college')
    assert get_response.status_code == 405

    post_response = admin_client.post(
        f'/tournament/{tournament_id}/activate/college',
        follow_redirects=False,
    )
    assert post_response.status_code == 302

    with app.app_context():
        tournament = Tournament.query.get(tournament_id)
        assert tournament.status == 'college_active'


def test_health_diag_requires_auth(app):
    client = app.test_client()
    response = client.get('/health/diag', follow_redirects=False)
    assert response.status_code == 302
    assert '/auth/login' in response.headers['Location']


def test_heat_generation_rejects_zero_max_stands(app):
    from database import db
    from models import Event, Tournament
    from models.competitor import ProCompetitor
    from services.heat_generator import generate_event_heats

    with app.app_context():
        tournament = Tournament(name='ZeroStand', year=2026, status='setup')
        db.session.add(tournament)
        db.session.flush()
        event = Event(
            tournament_id=tournament.id,
            name='Underhand',
            event_type='pro',
            gender='M',
            scoring_type='time',
            scoring_order='lowest_wins',
            stand_type='underhand',
            max_stands=0,
            status='pending',
        )
        db.session.add(event)
        db.session.flush()
        competitor = ProCompetitor(tournament_id=tournament.id, name='A', gender='M', status='active')
        competitor.set_events_entered([str(event.id)])
        db.session.add(competitor)
        db.session.commit()

        with pytest.raises(ValueError, match='invalid max_stands'):
            generate_event_heats(event)


def test_restore_rejects_schema_mismatch(app, admin_client):
    tournament_id = app.config['_FIX_TOURNAMENT_ID']
    fd, bad_restore = tempfile.mkstemp(
        prefix='remedial-restore-',
        suffix='.db',
        dir=app.instance_path,
    )
    os.close(fd)
    try:
        conn = sqlite3.connect(bad_restore)
        try:
            conn.execute('CREATE TABLE alembic_version (version_num VARCHAR(32) NOT NULL)')
            conn.execute("INSERT INTO alembic_version (version_num) VALUES ('old_revision')")
            conn.execute('CREATE TABLE tournaments (id INTEGER PRIMARY KEY)')
            conn.commit()
        finally:
            conn.close()

        with open(bad_restore, 'rb') as fh:
            response = admin_client.post(
                f'/reporting/{tournament_id}/restore',
                data={'backup_file': (io.BytesIO(fh.read()), 'restore.db')},
                content_type='multipart/form-data',
                follow_redirects=True,
            )
    finally:
        try:
            os.remove(bad_restore)
        except OSError:
            pass

    assert response.status_code == 200
    assert (
        b'missing required tables' in response.data
        or b'schema revision does not match' in response.data
    )


def test_export_results_creates_placeholder_sheet_for_empty_tournament(app):
    from database import db
    from models import Tournament
    from openpyxl import load_workbook
    from services.excel_io import export_results_to_excel

    fd, out = tempfile.mkstemp(
        prefix='remedial-export-',
        suffix='.xlsx',
        dir=app.instance_path,
    )
    os.close(fd)
    try:
        with app.app_context():
            tournament = Tournament(name='Empty Export', year=2026, status='setup')
            db.session.add(tournament)
            db.session.commit()

            export_results_to_excel(tournament, out)

        wb = load_workbook(out)
        assert 'Overview' in wb.sheetnames
    finally:
        try:
            os.remove(out)
        except OSError:
            pass
