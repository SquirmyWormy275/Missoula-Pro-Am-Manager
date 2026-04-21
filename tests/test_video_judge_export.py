"""
Tests for services/video_judge_export.py.

Covers row building + xlsx writing for the Video Judge workbook (2026-04-20).
Rules from docs/VIDEO_JUDGE_BRACKET_PLAN.md:
    - Skip scoring_type='bracket' (Birling).
    - Skip LIST_ONLY events (no heats).
    - Dual-run day-split events split into two sheets by run.
    - Triple-run events stack 3 rows per competitor on one sheet.
    - Hits/score events flip column labels to 'VJ Score 1 / VJ Score 2'.
    - Sheet names sanitised + deduped (31-char limit, no [ ] : * ? / \\).
    - Stable row order: heat / run / stand / competitor-name.

Run:  pytest tests/test_video_judge_export.py -v
"""

from __future__ import annotations

import openpyxl
import pytest

from services.video_judge_export import (
    VideoJudgeWorkbookError,
    _dedupe_sheet_names,
    _sanitize_sheet_name,
    build_video_judge_rows,
    write_workbook,
)
from tests.conftest import (
    make_college_competitor,
    make_event,
    make_pro_competitor,
    make_team,
    make_tournament,
)


def _make_heat(
    session, event, comp_ids, run_number=1, heat_number=1, stand_assignments=None
):
    """Create a Heat row with competitors + stand assignments."""
    from models import Heat

    h = Heat(
        event_id=event.id,
        heat_number=heat_number,
        run_number=run_number,
    )
    h.set_competitors(comp_ids)
    if stand_assignments:
        for cid, stand in stand_assignments.items():
            h.set_stand_assignment(cid, stand)
    session.add(h)
    session.flush()
    return h


# ---------------------------------------------------------------------------
# Sheet name sanitisation / dedup
# ---------------------------------------------------------------------------


class TestSheetNameHelpers:
    def test_sanitize_truncates_to_31_chars(self):
        long = "A" * 50
        out = _sanitize_sheet_name(long)
        assert len(out) == 31
        assert out == "A" * 31

    def test_sanitize_strips_invalid_chars(self):
        out = _sanitize_sheet_name("Men's [Underhand]: Round 1?/2*")
        for bad in "[]:*?/\\":
            assert bad not in out

    def test_sanitize_empty_becomes_sheet(self):
        assert _sanitize_sheet_name("") == "Sheet"
        assert _sanitize_sheet_name(None) == "Sheet"

    def test_dedupe_single_no_change(self):
        assert _dedupe_sheet_names(["Foo"]) == ["Foo"]

    def test_dedupe_two_same(self):
        assert _dedupe_sheet_names(["Foo", "Foo"]) == ["Foo", "Foo (2)"]

    def test_dedupe_three_same(self):
        assert _dedupe_sheet_names(["Foo", "Foo", "Foo"]) == [
            "Foo",
            "Foo (2)",
            "Foo (3)",
        ]


# ---------------------------------------------------------------------------
# build_video_judge_rows — event filtering + partner pairing
# ---------------------------------------------------------------------------


class TestBuildRows:
    def test_skips_bracket_events(self, db_session):
        t = make_tournament(db_session)
        birling = make_event(
            db_session,
            t,
            name="Birling",
            event_type="college",
            scoring_type="bracket",
            stand_type="birling",
            gender=None,
        )
        ev = make_event(
            db_session,
            t,
            name="Underhand Speed",
            event_type="college",
            scoring_type="time",
            stand_type="underhand",
            gender=None,
        )
        team = make_team(db_session, t)
        a = make_college_competitor(
            db_session,
            t,
            team,
            name="Alice Chopper",
            gender="F",
        )
        _make_heat(db_session, ev, [a.id], stand_assignments={a.id: 1})

        sheets = build_video_judge_rows(t)
        titles = list(sheets.keys())
        assert "Birling" not in titles
        assert "Underhand Speed" in titles

    def test_skips_list_only_events(self, db_session):
        t = make_tournament(db_session)
        # Axe Throw is LIST_ONLY by name.  Even with heats somehow, it should skip.
        axe = make_event(
            db_session,
            t,
            name="Axe Throw",
            event_type="college",
            scoring_type="score",
            stand_type="axe_throw",
            gender=None,
        )
        ev = make_event(
            db_session,
            t,
            name="Single Buck",
            event_type="college",
            scoring_type="time",
            stand_type="saw_hand",
            gender=None,
        )
        team = make_team(db_session, t)
        a = make_college_competitor(db_session, t, team, name="Alice", gender="F")
        _make_heat(db_session, axe, [a.id], stand_assignments={a.id: 1})
        _make_heat(db_session, ev, [a.id], stand_assignments={a.id: 1})

        sheets = build_video_judge_rows(t)
        assert "Axe Throw" not in sheets
        assert "Single Buck" in sheets

    def test_no_heats_event_excluded(self, db_session):
        t = make_tournament(db_session)
        ev = make_event(
            db_session,
            t,
            name="Underhand Speed",
            event_type="college",
            scoring_type="time",
            stand_type="underhand",
            gender=None,
        )
        # No heat — should not appear in workbook sheets.
        sheets = build_video_judge_rows(t)
        assert "Underhand Speed" not in sheets

    def test_partnered_event_one_row_per_pair(self, db_session):
        t = make_tournament(db_session)
        ev = make_event(
            db_session,
            t,
            name="Jack & Jill Sawing",
            event_type="pro",
            scoring_type="time",
            stand_type="saw_hand",
            gender=None,
            is_partnered=True,
        )
        a = make_pro_competitor(
            db_session,
            t,
            "Alice Chopper",
            gender="F",
            partners={"Jack & Jill Sawing": "Bob Splitter"},
        )
        b = make_pro_competitor(
            db_session,
            t,
            "Bob Splitter",
            gender="M",
            partners={"Jack & Jill Sawing": "Alice Chopper"},
        )
        _make_heat(
            db_session,
            ev,
            [a.id, b.id],
            stand_assignments={a.id: 1, b.id: 1},
        )

        sheets = build_video_judge_rows(t)
        entry = sheets["Jack & Jill Sawing"]
        assert len(entry["rows"]) == 1
        assert entry["rows"][0]["competitor_name"] == "Alice Chopper & Bob Splitter"

    def test_college_rows_include_team_code(self, db_session):
        t = make_tournament(db_session)
        team = make_team(db_session, t, code="UM-A", school="University of Montana")
        ev = make_event(
            db_session,
            t,
            name="Underhand Speed",
            event_type="college",
            scoring_type="time",
            stand_type="underhand",
            gender=None,
        )
        c = make_college_competitor(
            db_session,
            t,
            team,
            name="Alice Chopper",
            gender="F",
        )
        _make_heat(db_session, ev, [c.id], stand_assignments={c.id: 1})

        sheets = build_video_judge_rows(t)
        row = sheets["Underhand Speed"]["rows"][0]
        assert row["team_code"] == "UM-A"

    def test_pro_rows_team_code_none(self, db_session):
        t = make_tournament(db_session)
        ev = make_event(
            db_session,
            t,
            name="Springboard",
            event_type="pro",
            scoring_type="time",
            stand_type="springboard",
        )
        c = make_pro_competitor(db_session, t, "Alice Chopper", gender="F")
        _make_heat(db_session, ev, [c.id], stand_assignments={c.id: 1})

        sheets = build_video_judge_rows(t)
        row = sheets["Springboard"]["rows"][0]
        assert row["team_code"] is None

    def test_day_split_dual_run_splits_into_two_sheets(self, db_session):
        """Speed Climb is day-split dual-run → 'Speed Climb - Run 1' and 'Run 2'."""
        t = make_tournament(db_session)
        ev = make_event(
            db_session,
            t,
            name="Speed Climb",
            event_type="college",
            scoring_type="time",
            stand_type="speed_climb",
            gender=None,
            requires_dual_runs=True,
        )
        team = make_team(db_session, t)
        a = make_college_competitor(db_session, t, team, name="Alice", gender="F")
        _make_heat(db_session, ev, [a.id], run_number=1, stand_assignments={a.id: 1})
        _make_heat(
            db_session,
            ev,
            [a.id],
            run_number=2,
            heat_number=1,
            stand_assignments={a.id: 1},
        )

        sheets = build_video_judge_rows(t)
        titles = list(sheets.keys())
        # Speed Climb is in DAY_SPLIT_EVENT_NAMES so we expect separate sheets.
        assert "Speed Climb - Run 1" in titles
        assert "Speed Climb - Run 2" in titles
        assert "Speed Climb" not in titles

    def test_stable_row_ordering(self, db_session):
        t = make_tournament(db_session)
        ev = make_event(
            db_session,
            t,
            name="Underhand",
            event_type="pro",
            scoring_type="time",
            stand_type="underhand",
            gender=None,
        )
        a = make_pro_competitor(db_session, t, "Zach Z", gender="M")
        b = make_pro_competitor(db_session, t, "Anna A", gender="M")
        c = make_pro_competitor(db_session, t, "Bob B", gender="M")
        # Two heats, out-of-order.
        _make_heat(
            db_session,
            ev,
            [a.id, b.id],
            heat_number=2,
            stand_assignments={a.id: 2, b.id: 1},
        )
        _make_heat(
            db_session,
            ev,
            [c.id],
            heat_number=1,
            stand_assignments={c.id: 1},
        )
        sheets = build_video_judge_rows(t)
        rows = sheets["Underhand"]["rows"]
        # Ordered by heat_number, then stand_number.  Heat 1 first, then heat 2.
        assert rows[0]["heat_number"] == 1 and rows[0]["competitor_name"] == "Bob B"
        # Heat 2: stand 1 (Anna A) before stand 2 (Zach Z).
        assert rows[1]["heat_number"] == 2 and rows[1]["competitor_name"] == "Anna A"
        assert rows[2]["heat_number"] == 2 and rows[2]["competitor_name"] == "Zach Z"


# ---------------------------------------------------------------------------
# write_workbook — xlsx output contents
# ---------------------------------------------------------------------------


class TestWriteWorkbook:
    def _read_sheet_names(self, path):
        wb = openpyxl.load_workbook(path)
        return wb.sheetnames

    def test_empty_sheets_writes_placeholder_workbook(self, db_session, tmp_path):
        from collections import OrderedDict

        path = str(tmp_path / "out.xlsx")
        write_workbook(OrderedDict(), path)
        names = self._read_sheet_names(path)
        assert names == ["No Events"]

    def test_writes_sheet_per_event(self, db_session, tmp_path):
        t = make_tournament(db_session)
        ev1 = make_event(
            db_session,
            t,
            name="Underhand Speed",
            event_type="college",
            scoring_type="time",
            stand_type="underhand",
            gender=None,
        )
        ev2 = make_event(
            db_session,
            t,
            name="Springboard",
            event_type="pro",
            scoring_type="time",
            stand_type="springboard",
        )
        team = make_team(db_session, t)
        a = make_college_competitor(db_session, t, team, name="Alice", gender="F")
        b = make_pro_competitor(db_session, t, "Bob", gender="M")
        _make_heat(db_session, ev1, [a.id], stand_assignments={a.id: 1})
        _make_heat(db_session, ev2, [b.id], stand_assignments={b.id: 1})

        sheets = build_video_judge_rows(t)
        path = str(tmp_path / "vj.xlsx")
        write_workbook(sheets, path)
        names = self._read_sheet_names(path)
        assert "Underhand Speed" in names
        assert "Springboard" in names

    def test_timed_event_uses_timer_column_headers(self, db_session, tmp_path):
        t = make_tournament(db_session)
        ev = make_event(
            db_session,
            t,
            name="Springboard",
            event_type="pro",
            scoring_type="time",
            stand_type="springboard",
        )
        b = make_pro_competitor(db_session, t, "Bob", gender="M")
        _make_heat(db_session, ev, [b.id], stand_assignments={b.id: 1})

        sheets = build_video_judge_rows(t)
        path = str(tmp_path / "vj.xlsx")
        write_workbook(sheets, path)

        wb = openpyxl.load_workbook(path)
        ws = wb["Springboard"]
        header_row = [c.value for c in ws[1]]
        assert "VJ Timer 1" in header_row
        assert "VJ Timer 2" in header_row
        assert "VJ Score 1" not in header_row

    def test_hits_event_uses_score_column_headers(self, db_session, tmp_path):
        t = make_tournament(db_session)
        ev = make_event(
            db_session,
            t,
            name="Underhand Hard Hit",
            event_type="college",
            scoring_type="hits",
            stand_type="underhand",
            gender=None,
        )
        team = make_team(db_session, t)
        a = make_college_competitor(db_session, t, team, name="Alice", gender="F")
        _make_heat(db_session, ev, [a.id], stand_assignments={a.id: 1})

        sheets = build_video_judge_rows(t)
        path = str(tmp_path / "vj.xlsx")
        write_workbook(sheets, path)

        wb = openpyxl.load_workbook(path)
        ws = wb["Underhand Hard Hit"]
        header_row = [c.value for c in ws[1]]
        assert "VJ Score 1" in header_row
        assert "VJ Score 2" in header_row
        assert "VJ Timer 1" not in header_row

    def test_sheet_name_truncation_does_not_crash(self, db_session, tmp_path):
        from collections import OrderedDict
        from types import SimpleNamespace

        long_title = "A Very Long Event Title That Exceeds Thirty One Characters"
        assert len(long_title) > 31
        sheets = OrderedDict()
        sheets[long_title] = {
            "event": SimpleNamespace(scoring_type="time", event_type="pro"),
            "rows": [],
        }
        path = str(tmp_path / "vj.xlsx")
        write_workbook(sheets, path)
        names = self._read_sheet_names(path)
        assert len(names[0]) == 31
        assert names[0] == long_title[:31]

    def test_invalid_chars_stripped_from_sheet_name(self, db_session, tmp_path):
        from collections import OrderedDict
        from types import SimpleNamespace

        title = "Men's [Underhand]: Round 1?/2*"
        sheets = OrderedDict()
        sheets[title] = {
            "event": SimpleNamespace(scoring_type="time", event_type="pro"),
            "rows": [],
        }
        path = str(tmp_path / "vj.xlsx")
        write_workbook(sheets, path)
        names = self._read_sheet_names(path)
        for bad in "[]:*?/\\":
            assert bad not in names[0]

    def test_openpyxl_failure_raises_custom_exception(self, tmp_path):
        """An unwritable path → VideoJudgeWorkbookError, not raw OSError/500.

        Openpyxl is imported lazily inside write_workbook so we can't cheaply
        monkeypatch it.  Instead, point at a non-existent nested directory
        so openpyxl's save() fails; write_workbook must catch and re-raise.
        """
        from collections import OrderedDict

        bad_path = str(tmp_path / "does" / "not" / "exist" / "vj.xlsx")
        with pytest.raises(VideoJudgeWorkbookError):
            write_workbook(OrderedDict(), bad_path)
